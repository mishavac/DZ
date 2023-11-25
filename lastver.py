import re
import pandas as pd
from openpyxl.utils import get_column_letter

# Загрузите файл Excel в DataFrame
df = pd.read_excel(r'C:\Users\NOVIKOV\Desktop\ebala\raspisanie.xlsx', sheet_name='16.10.2023-21.10.2023', header=None)

# Ищем колонку с названием "Группа:"
column_name = 'Группа:'
group_row = None

for index, row in df.iterrows():
    for col in df.columns:
        cell_value = row[col]
        if isinstance(cell_value, str) and column_name.lower() in cell_value.lower():
            print(f'Найдено в строке {index + 1}, колонка {col + 1}')
            group_row = row
            break

    if group_row is not None:
        break

if group_row is not None:
    print("Найденная строка:")
    data = []
    for i, value in enumerate(group_row):
        column_letter = get_column_letter(i + 1)
        data.append(f'Строка {index + 1}, Колонка {column_letter}: {value}')

    # Удаляем строки с NaN
    data = [line for line in data if 'nan' not in str(line).lower()]

    # Получаем индекс колонки, начиная с которой нужно вывести столбец
    start_column_index = group_row.index[0]

    # Выводим все столбцы начиная с найденной строки
    result_data = []
    for j in range(len(df.columns[start_column_index:])):
        column_data = []
        for i, value in df.iloc[index:, start_column_index + j].items():
            column_letter = get_column_letter(start_column_index + j + 1)
            if not pd.isna(value):
                column_data.append(f'Строка {i + 1}, Колонка {column_letter}: {value}')
        
        result_data.append('\n'.join(column_data))

    # Сохраняем результат в CSV файл
    with open('output_result.csv', 'w', newline='', encoding='utf-8') as csvfile:
        csvfile.write('\n'.join(result_data))